from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


webpage = 'https://www.coinranking.com'


page = urlopen(webpage)        


soup = BeautifulSoup(page, 'html.parser')


title = soup.title


print(title.text)


# wb = xl.Workbook()
# ws = wb.active


# ws.title = 'Cryptocurrency Report'


# ws['A1'] = 'Name'
# ws['B1'] = 'Price'
# ws['C1'] = 'Percent Change'


coin_rows = soup.findAll("tr")


for x in range(4,9):
   td = coin_rows[x].findAll('td')
   print(td[1].text)
   # name = td[0].text
   # price = td[1].text
   # change = td[3].text


   # ws['A' + str(x-2)] = name
   # ws['B' + str(x-2)] = price
   # ws['C' + str(x-2)] = change


# ws.column_dimensions['A'].width = 10
# ws.column_dimensions['B'].width = 15
# ws.column_dimensions['A'].width = 10


# header_font = Font(size=16, bold=True)


# for cell in ws[1:1]:
#     cell.font = header_font


# wb.save("Cryptocurrency Report.xlsx")

